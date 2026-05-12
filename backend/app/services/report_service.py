from datetime import datetime
from io import BytesIO, StringIO
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


BRAND_NAVY = "#0f2440"
BRAND_BLUE = "#2f80ed"
BRAND_GREEN = "#22a06b"
BRAND_MIST = "#eef7fb"
TEXT_MUTED = "#64748b"


def _pretty(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return f"{value:,.2f}" if isinstance(value, float) else f"{value:,}"
    return str(value)


def _label(column: str) -> str:
    labels = {
        "created_at": "Creado",
        "fecha_pago": "Fecha de pago",
        "fecha": "Fecha",
        "nombre": "Nombre",
        "email": "Correo",
        "telefono": "Teléfono",
        "direccion": "Dirección",
        "estado": "Estado",
        "tipo": "Tipo",
        "categoria": "Categoría",
        "descripcion": "Descripción",
        "monto": "Monto",
        "titulo": "Título",
        "prioridad": "Prioridad",
        "lugar": "Lugar",
    }
    return labels.get(column, column.replace("_", " ").capitalize())


class BaseReport:
    title = "Reporte"
    subtitle = "Reporte comunitario generado desde Neighbord"
    columns: list[str] = []

    def __init__(self, rows: list[dict]):
        self.rows = rows

    def _table_rows(self) -> list[list[str]]:
        return [[_label(col) for col in self.columns]] + [
            [_pretty(row.get(col, ""))[:80] for col in self.columns]
            for row in self.rows
        ]

    def to_pdf(self) -> bytes:
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=0.45 * inch,
            leftMargin=0.45 * inch,
            topMargin=0.4 * inch,
            bottomMargin=0.45 * inch,
        )
        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        title_style.textColor = colors.HexColor(BRAND_NAVY)
        title_style.fontName = "Helvetica-Bold"
        title_style.fontSize = 22
        subtitle_style = styles["BodyText"]
        subtitle_style.textColor = colors.HexColor(TEXT_MUTED)
        subtitle_style.fontSize = 9

        story = [
            Paragraph("Neighbord", styles["Heading2"]),
            Paragraph(self.title, title_style),
            Paragraph(f"{self.subtitle} · Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} · Registros: {len(self.rows)}", subtitle_style),
            Spacer(1, 0.18 * inch),
        ]

        data = self._table_rows()
        usable_width = 10.1 * inch
        col_width = usable_width / max(len(self.columns), 1)
        table = Table(data, repeatRows=1, colWidths=[col_width] * len(self.columns))
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_NAVY)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 9),
            ("TOPPADDING", (0, 0), (-1, 0), 9),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(BRAND_MIST)]),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor(BRAND_NAVY)),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dbe6ee")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        return buffer.read()

    def to_csv(self) -> bytes:
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=self.columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(self.rows)
        return output.getvalue().encode("utf-8-sig")

    def to_xlsx(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.title[:31]

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(self.columns), 1))
        sheet["A1"] = "Neighbord"
        sheet["A1"].font = Font(bold=True, size=18, color=BRAND_NAVY.replace("#", ""))
        sheet["A1"].alignment = Alignment(horizontal="left")

        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(self.columns), 1))
        sheet["A2"] = f"{self.title} · Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} · Registros: {len(self.rows)}"
        sheet["A2"].font = Font(size=10, color=TEXT_MUTED.replace("#", ""))

        header_fill = PatternFill("solid", fgColor=BRAND_NAVY.replace("#", ""))
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="DBE6EE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_index, column in enumerate(self.columns, start=1):
            cell = sheet.cell(row=4, column=col_index, value=_label(column))
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        for row_index, row in enumerate(self.rows, start=5):
            fill = PatternFill("solid", fgColor="EEF7FB" if row_index % 2 == 0 else "FFFFFF")
            for col_index, column in enumerate(self.columns, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=row.get(column, ""))
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_index, column in enumerate(self.columns, start=1):
            values = [_label(column), *[_pretty(row.get(column, "")) for row in self.rows[:80]]]
            width = min(36, max(12, max(len(str(value)) for value in values) + 2))
            sheet.column_dimensions[get_column_letter(col_index)].width = width

        sheet.freeze_panes = "A5"
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()


class VecinosReport(BaseReport):
    title = "Reporte de vecinos"
    subtitle = "Padrón comunitario y estado de cuentas"
    columns = ["nombre", "email", "telefono", "direccion", "estado"]


class FinancieroReport(BaseReport):
    title = "Reporte financiero"
    subtitle = "Movimientos financieros registrados"
    columns = ["tipo", "categoria", "descripcion", "monto", "fecha"]


class SolicitudesReport(BaseReport):
    title = "Reporte de solicitudes"
    subtitle = "Seguimiento de reclamos y solicitudes vecinales"
    columns = ["titulo", "categoria", "prioridad", "estado", "created_at"]


class ActaReunionReport(BaseReport):
    title = "Acta de reuniones"
    subtitle = "Reuniones y asambleas registradas"
    columns = ["titulo", "fecha", "lugar", "estado"]


class CronogramaReport(BaseReport):
    title = "Cronograma comunitario"
    subtitle = "Planificación de actividades y reuniones"
    columns = ["titulo", "descripcion", "fecha", "estado"]


class ReportFactory:
    reports = {
        "vecinos": VecinosReport,
        "financiero": FinancieroReport,
        "solicitudes": SolicitudesReport,
        "actas": ActaReunionReport,
        "cronograma": CronogramaReport,
    }

    @classmethod
    def create(cls, report_type: str, rows: list[dict]) -> BaseReport:
        report_cls = cls.reports.get(report_type)
        if not report_cls:
            raise ValueError("Tipo de reporte no soportado")
        return report_cls(rows)
